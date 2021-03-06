# coding:utf-8
import smtplib
import os.path as pth
import time, os
from email import encoders
from email.mime.text import MIMEText
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from basic_info.setting import email_to
from smtplib import SMTP_SSL
from openpyxl import load_workbook
from api_test_cases.get_execution_output_json import abs_dir, GetCheckoutDataSet
from new_api_cases.check_result import table_dir

# abs_dir2 = lambda n: os.path.abspath(os.path.join(os.path.dirname(__file__), n))


def sendEmail(content, title, from_name, from_address, to_address, serverport, serverip, username, password):
    # 邮件对象:
    msg = MIMEMultipart()
    msg['Subject'] = Header(title, 'utf-8')
    # 这里的to_address只用于显示，必须是一个string
    msg['To'] = ','.join(to_address)
    msg['From'] = from_name
    # 邮件正文是MIMEText:
    msg.attach(MIMEText(content, 'html', 'utf-8'))
    # 添加附件就是加上一个MIMEBase，从本地读取一个文件:
    with open('E:\Reports\Test_report.html', 'rb') as f:
        # 设置附件的MIME和文件名，这里是png类型:
        mime = MIMEBase('report', 'html', filename='Test_Report.html')
        # mime = MIMEBase('report', 'html', filename=filename)
        # 加上必要的头信息:
        mime.add_header('Content-Disposition', 'attachment', filename='Test_Report.html')
        mime.add_header('Content-ID', '<0>')
        mime.add_header('X-Attachment-Id', '0')
        # 把附件的内容读进来:
        mime.set_payload(f.read())
        # 用Base64编码:
        encoders.encode_base64(mime)
        # 添加到MIMEMultipart:
        msg.attach(mime)

    try:
        s = smtplib.SMTP_SSL(serverip, serverport)
        s.login(username, password)
        # 这里的to_address是真正需要发送的到的mail邮箱地址需要的是一个list
        s.sendmail(from_address, to_address, msg.as_string())
        print('%s----发送邮件成功' % time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    except Exception as err:
        print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        print(err)
        #HEFEN_D = pth.abspath(pth.dirname(__file__))

def main2():
    # from run import filename
    TO = [email_to["gubingjie"], email_to["daming"]]
    # TO = [email_to["gubingjie"]]
    config = {
    "from": "ruifan_test@163.com",
    "from_name": '189环境自动化测试报告',
    "to": TO,
    "serverip": "smtp.163.com",
    "serverport": "465",
    "username": "ruifan_test@163.com",
    "password": "ruifantest2018"  # QQ邮箱的SMTP授权码
             }

    title = "自动化测试报告"
    f = open("E:\Reports\Test_Report.html", 'rb')
    mail_body = f.read()
    f.close()
    sendEmail(mail_body, title, config['from_name'], config['from'], config['to'], config['serverport'], config['serverip'], config['username'], config['password'])


# 使用该方法发送邮件
def main3(report_path):
    # 163邮箱smtp服务器
    host_server = "smtp.163.com"
    # sender_163为发件人的163邮箱
    sender_163 = "ruifan_test@163.com"
    # pwd为163邮箱的授权码
    pwd = "ruifantest2018"
    # 发件人的邮箱
    sender_163_mail = "ruifan_test@163.com"
    # 收件人邮箱
    receivers = ['bingjie.gu@inforefiner.com', 'zhiming.wang@inforefiner.com', 'yuan.peng@inforefiner.com', 'anchong.wang@inforefiner.com', 'qian.feng@inforefiner.com'] # 定时任务使用
    # receivers = ['bingjie.gu@inforefiner.com']  # 调试使用
    msg = MIMEMultipart()
    # 邮件的正文内容----flow执行结果
    f = load_workbook(abs_dir("flow_dataset_info.xlsx"))
    f_sheet = f.get_sheet_by_name("flow_info")
    cols = f_sheet.max_column
    rows = f_sheet.max_row
    succeed = 0
    succeed_flow = []
    failed = 0
    failed_flow = []
    flow_failed_detail = []
    flow_id_list = GetCheckoutDataSet().file_flowid_count()
    total = len(flow_id_list)
    print('flow总数total:', total)
    detail_msg = ''' '''
    for row in range(2, rows+1):
        if f_sheet.cell(row=row, column=9).value == "fail":
            detail_msg += f_sheet.cell(row=row, column=10).value + '\n'
            if f_sheet.cell(row=row, column=3).value:
                failed_flow.append(f_sheet.cell(row=row, column=3).value)
            else:
                for i in range(row, 2, -1):
                    if f_sheet.cell(row=i-1, column=3).value:
                        failed_flow.append(f_sheet.cell(row=i-1, column=3).value)
                        break
        elif f_sheet.cell(row=row, column=9).value == "pass":
            if f_sheet.cell(row=row, column=3).value:
                succeed_flow.append(f_sheet.cell(row=row, column=3).value)
                # print(succeed_flow)
            else:
                for i in range(row, 2, -1):
                    if f_sheet.cell(row=i-1, column=3).value:
                        succeed_flow.append(f_sheet.cell(row=i-1, column=3).value)
                        break
    failed_flow_s = list(set(failed_flow))
    succeed_flow_s = list(set(succeed_flow))
    for disct_id in (disct_ids for disct_ids in failed_flow_s if disct_ids in succeed_flow_s):
        succeed_flow_s.remove(disct_id)
    # 邮件的正文内容----API执行结果
    # 统计api执行结果，加入到邮件正文中，失败的用例name，失败的原因
    api_cases_table = load_workbook(table_dir('api_cases.xlsx'))
    cases_sheet = api_cases_table.get_sheet_by_name('tester')
    sheet_rows = cases_sheet.max_row
    cases_num = sheet_rows - 1
    pass_cases = 0
    failed_cases = 0
    failed_cases_list = []
    for row in range(2, sheet_rows+1):
        if cases_sheet.cell(row=row, column=14).value == 'pass':
            pass_cases += 1
        elif cases_sheet.cell(row=row, column=14).value == 'fail':
            failed_cases += 1
            # k = cases_sheet.cell(row=row, column=2).value
            # v = cases_sheet.cell(row=row, column=15).value
            # dict = {k: v}
            failed_cases_list.append(cases_sheet.cell(row=row, column=15).value)
        else:
            print('请确认第%d行测试用例测试结果' % row)


    # 邮件的正文内容
    filename = time.strftime("%Y%m%d%H", time.localtime()) + '_report.html'
    if len(failed_flow_s) > 0:
        mail_content = '\n各位好:'+'\n' + '\n' + \
                    '一、API的测试用例结果请参考附件<<%s>>' % filename + '\n'\
                    + '-------------------------------------------------------------' + '\n'\
                       + '二、Flow用例共 %d 个\n成功%d个, 失败 %d个\n失败的flow name为 %s\n失败原因为：\n ' % (total, len(succeed_flow_s), len(failed_flow_s), failed_flow_s) + detail_msg + \
                       '-------------------------------------------------------------' + '\n' + \
                       '三、 API测试用例共%d个，\n成功%d个， 失败%d个,失败的case info： \n%s' \
                       % (cases_num, pass_cases, failed_cases, failed_cases_list)
    else:
        mail_content = '\n各位好:' + '\n' + \
                       '一、API的测试用例结果请参考附件<<%s>>' % filename + '\n' \
                   + '--------------------------------------------------------------' + '\n' \
                       + '二、flow用例共 %d 个\n成功%d个, 失败 %d个\n' % (total, len(succeed_flow_s), len(failed_flow_s)) + '\n' \
                       + '--------------------------------------------------------------' + '\n' + \
                       '三、API测试用例共%d个，\n成功%d个， 失败%d个,失败的case info: \n%s' \
                       % (cases_num, pass_cases, failed_cases, failed_cases_list)
    # print(mail_content)
    # 邮件标题
    # mail_title = time.strftime("%Y-%m-%d 自动化日报", time.localtime())
    mail_title = time.strftime("%Y-%m-%d", time.localtime()) + ' API自动化测试日报'
    # 添加邮件正文，格式 MIMEText:
    msg.attach(MIMEText(mail_content, "plain", 'utf-8'))

    # 添加附件，就是加上一个MIMEBase，从本地读取一个文件:
    # 添加API用例集执行报告
    filename = time.strftime("%Y%m%d%H", time.localtime()) + '_report.html'
    with open(report_path, 'rb') as f:
        # 设置附件的MIME和文件名，这里是html类型:
        mime = MIMEBase('report', 'html', filename=filename)
        # 加上必要的头信息:
        mime.add_header('Content-Disposition', 'attachment', filename=filename)
        mime.add_header('Content-ID', '<0>')
        mime.add_header('X-Attachment-Id', '0')
        # 把附件的内容读进来:
        mime.set_payload(f.read())
        # # 用Base64编码:
        encoders.encode_base64(mime)
        # 添加到MIMEMultipart:
        msg.attach(mime)
    # 添加api用例 excel表格
    apicases_filepath = table_dir('api_cases.xlsx')
    with open(apicases_filepath, 'rb') as a:
        # 设置附件的MIME和文件名，这里是html类型:
        mime = MIMEBase('report', 'xlsx', filename='api_cases.xlsx')
        # 加上必要的头信息:
        mime.add_header('Content-Disposition', 'attachment', filename='api_casex.xlsx')
        mime.add_header('Content-ID', '<0>')
        mime.add_header('X-Attachment-Id', '0')
        # 把附件的内容读进来:
        mime.set_payload(a.read())
        # # 用Base64编码:
        encoders.encode_base64(mime)
        # 添加到MIMEMultipart:
        msg.attach(mime)

    flow_filepath = abs_dir('flow_dataset_info.xlsx')
    with open(flow_filepath, 'rb') as ff:
        # 设置附件的MIME和文件名，这里是html类型:
        mime = MIMEBase('report', 'xlsx', filename='flow_info.xlsx')
        # 加上必要的头信息:
        mime.add_header('Content-Disposition', 'attachment', filename='flow_info.xlsx')
        mime.add_header('Content-ID', '<0>')
        mime.add_header('X-Attachment-Id', '0')
        # 把附件的内容读进来:
        mime.set_payload(ff.read())
        # # 用Base64编码:
        encoders.encode_base64(mime)
        # 添加到MIMEMultipart:
        msg.attach(mime)

    # ssl登录
    smtp = SMTP_SSL(host_server)
    # set_debuglevel()是用来调试的。参数值为1表示开启调试模式，参数值为0关闭调试模式
    smtp.set_debuglevel(0)
    smtp.ehlo(host_server)
    smtp.login(sender_163, pwd)
    msg["Subject"] = Header(mail_title, 'utf-8')
    msg["From"] = sender_163_mail
    msg["To"] = Header("顾冰洁，王志明，彭媛", 'utf-8')  # 接收者的别名
    smtp.sendmail(sender_163_mail, receivers, msg.as_string())
    print('%s----发送邮件成功' % time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    smtp.quit()

# report_path = 'E:\Reports\\2018122813_report.html'
# main3(report_path)
# apicases_filepath = table_dir('api_cases.xlse')
# # # print(apicases_filepath)