import xlrd, xlwt
from openpyxl import workbook
from openpyxl import load_workbook
from xlutils.copy import copy


# 设置表格样式
def set_style(name, height, bold=False):
    style = xlwt.XFStyle()
    font = xlwt.Font()
    font.name = name
    font.bold = bold
    font.color_index = 4
    font.height = height
    style.font = font
    return style


# 写Excel
def write_excel():
    f = xlwt.Workbook()
    sheet1 = f.add_sheet('flow_info', cell_overwrite_ok=True)
    row0 = ["id", "flow_id", "except_result", "actual", "test_result"]
    colum0 = ["35033c8d-fadc-4628-abf9-6803953fba34"]
    # 写第一行
    for i in range(0, len(row0)):
        sheet1.write(0, i, row0[i], set_style('Times New Roman', 220, True))
    # 写第二行的第二列
    for i in range(0, len(colum0)):
        sheet1.write(i + 1, 0, i + 1, set_style("Times New Roman", 220, True))
        sheet1.write(i + 1, 1, colum0[i], set_style('Times New Roman', 220, True))

    sheet1.write(1, 2, 'data_json')
    f.save('flow_dataset_info.xls')


# 读
def read_execl():
    data = xlrd.open_workbook("flow_dataset_info.xls")
    table1 = data.sheets()[0]
    table2 = data.sheet_by_name("flow_info")
    data1 = table1.ncols
    data2 = table1.nrows
    data3 = table1.cell(0, 2).value
    print(table1)
    return data1, data2, data3


# 改
def update_excel():
    wb = xlrd.open_workbook("flow_dataset_info.xls")  # 打开表
    crb = copy(wb)  # 复制表
    crb_sheet = crb.get_sheet(0)  # 打开复制表的第一个表单
    # crb_sheet.write(1, 3, "") # 第2行第四列，修改为“”

    crb_sheet.write(1, 3, )
    crb.save("flow_dataset_info.xls")


def check():
    table = xlrd.open_workbook("flow_dataset_info.xls")
    table_sheet = table.sheets()[0]
    c_table = copy(table)
    c_table_sheet = c_table.get_sheet(0)

    c_rows = table_sheet.nrows

    print('行数：', c_rows)
    # print(table_sheet.cell(1, 2).value)
    for i in range(1, c_rows):
        # print(table_sheet.cell(i, 2).value)
        if table_sheet.cell(i, 2).value:
            if table_sheet.cell(i, 2).value == table_sheet.cell(i, 3).value:
                c_table_sheet.write(i, 4, "pass")
            else:
                c_table_sheet.write(i, 4, "fail")

    c_table.save("flow_dataset_info.xls")
def openpyxl_read():
    wb = load_workbook('test.xlsx')
    sheetnames = wb.get_sheet_names()
    flow_info = wb.get_sheet_by_name(sheetnames[0])
    rows = flow_info.max_row
    columns = flow_info.max_column
    flow_info.cell(row=2,column=12,value=8888)
    print(flow_info.cell(row=1, column=3).value, rows, columns)
    print(flow_info.cell(row=2, column=12).value)

if __name__ == '__main__':
    print(openpyxl_read())